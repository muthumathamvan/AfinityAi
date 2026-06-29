from django.shortcuts import render
from django.shortcuts import render
from openpyxl import load_workbook
from openpyxl import Workbook
import cv2
import numpy as np
from django.conf import settings
from django.http import JsonResponse
from twilio.rest import Client
from nltk.sentiment import SentimentIntensityAnalyzer
import os
import re
import json
import pandas as pd
import httpx
from django.conf import settings
from django.http import HttpResponse
import time
from .models import *
from datetime import datetime
from pathlib import Path
import requests
from io import BytesIO
import math
from django.core.files.base import ContentFile

from elevenlabs import ElevenLabs, OutboundCallRecipient
from django.utils.timezone import now
from django.shortcuts import redirect



# Create your views here.
def home(requests):
    today = now().date()
    print(today)
    Total_count = call_details.objects.filter(phone_no__isnull=False).count()
    # print(Total_count)
    Interested_count = call_details.objects.filter(feedback="**Interested**", phone_no__isnull=False).count()
    # print(Interested_count)
    Not_pick_count = call_details.objects.filter(feedback="**Not pick call**", phone_no__isnull=False).count()
    today_count = call_details.objects.filter(date = today, phone_no__isnull=False).count()
    data = call_details.objects.filter(feedback="**Interested**").order_by('-date')[:5]
    return render(requests, "index.html", {'data':data, 'Total_count':Total_count, 'Interested_count':Interested_count, 'Not_pick_count':Not_pick_count, 'today_count':today_count, 'name':'Mathavan'})

def ai_form(requests):
    return render(requests, "ai-calling.html")

def login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        if username=="admin" and password=="admin@123":
            data = call_details.objects.filter(feedback="**Interested**").order_by('-date')[:5]
            return redirect('home')
        return render(requests, "auth-login.html",{'message':"invalid username or password"})
    return render(request, "auth-login.html")

def logout(request):
    return render(request, "auth-login.html")

def call_process(request):
    if request.method == 'POST':
        Name = request.POST['name']
        Time = request.POST['time']
        Date = request.POST['date']
        try:
            excel_file = request.FILES['file']
            wb = load_workbook(excel_file)
            ws = wb.active
            call_list =[]
            for row in ws.iter_rows(min_row=2, values_only=True):
                print(row[0], row[1], row[2], row[3])
                ai_call = generate_speech(row[0], row[1], row[2]) 
            
                call_list.append(ai_call)
            time.sleep(80)
            call_records(call_list, Name)
            return render(request, 'ai-calling.html')
        except:
            excel_file = request.POST['number']
            ai_call =  generate_speech_single(excel_file)
        
            time.sleep(100)
            call_records_single(ai_call, Name)
            return render(request, 'ai-calling.html')
    
    return render(request, "ai-calling.html")

# Ai call make (bulk call)
def generate_speech(name, address, to_number):
    try:
        client = ElevenLabs(api_key=settings.ELEVENLABS_API_KEY)
        phone_numbers = client.conversational_ai.phone_numbers.list()

        for number in phone_numbers:
            print(vars(number))

        response = client.conversational_ai.twilio.outbound_call(
                    agent_id= settings.ELEVENLABS_AGENTS,
                    agent_phone_number_id=settings.AGENT_PHONE_NUMBER_ID,
                    to_number= f"+91{to_number}",
                )
        
        
        data ={'conversation_id':response.conversation_id,
               'name':name,
               'address':address,
               }
        return data

    except Exception as e:
        return JsonResponse({"error": "Error while scheduling call", "details": str(e)}, status=500)

# Ai call make (single call)
def generate_speech_single(to_number):
    try:
        # to_number = 9600388948
        # to_number = 9600388948
        client = ElevenLabs(api_key=settings.ELEVENLABS_API_KEY)
        phone_numbers = client.conversational_ai.phone_numbers.list()

        for number in phone_numbers:
            print(vars(number))

        response = client.conversational_ai.twilio.outbound_call(
                    agent_id=settings.ELEVENLABS_AGENTS,
                    agent_phone_number_id= settings.AGENT_PHONE_NUMBER_ID,
                    to_number= f"+91{to_number}",
                )
        
        
        data ={'conversation_id':response.conversation_id
               }
        return data
        # return redirect('home')
        

    except Exception as e:
        return JsonResponse({"error": "Error while scheduling call", "details": str(e)}, status=500)

# Ai call records (bulk call)
def call_records(data, uploder_name):

    client = ElevenLabs(api_key=settings.ELEVENLABS_API_KEY)

    all_results = []
    for list_data in data:
        conversation_id = list_data['conversation_id']
        name = list_data['name']
        address = list_data['address']
       
        conversation = client.conversational_ai.conversations.get(conversation_id)
        conversation_json = json.dumps(conversation.__dict__, default=str)
        conversation_dict = json.loads(conversation_json)

        print(conversation)
        analysis = conversation_dict.get("analysis", "")
        analysis1 = conversation_dict.get("conversation_initiation_client_data", "")

        # Extract called number
        match1 = re.search(r"'system__called_number':\s*'([^']+)'", analysis1)
        match2 = re.search(r"external_number='(\+\d+)'", analysis1)
        match2 = re.search(r"external_number='(\+\d+)'", conversation_dict["metadata"])
        external_number = match2.group(1) if match2 else None
        called_number = match1.group(1) if match1 else external_number

        
        # Extract transcript summary
        try:
            match = re.search(r"transcript_summary='(.*?)'", analysis)
            transcript_summary = match.group(1).encode().decode('unicode_escape')
            feedback = classify_interest(transcript_summary)
            Audio = audio(conversation_id)
            audio_filename = Audio[0]['audio_filename']
            response_content = Audio[0]['response_content']

        except Exception as e:
            print(f"Error extracting called_number: {e}")
            feedback = "**Not pick call**"
            transcript_summary = ""
            audio_filename = "NO"
            # audio_filename = ""
            response_content =""

        # Store result
        all_results.append({
            "Name": name,
            "Address": address,
            "Called Number": called_number,
            "Feedback": feedback,
            "Transcript Summary": transcript_summary,
            "audio_filename":audio_filename,
            "response_content":response_content
            
        })
    for result in all_results:
        instance = call_details(
            uploader_name=uploder_name,
            name=result["Name"],
            address=result["Address"],
            phone_no=result["Called Number"],
            feedback=result["Feedback"],
            summary=result["Transcript Summary"],
            audio_flie=result["audio_filename"],
            time=datetime.now()
        )
        
        if result["response_content"] and result["audio_filename"].endswith('.mp3'):
            instance.audio_url.save(
                result["audio_filename"],
                ContentFile(result["response_content"]),
                save=False
            )

        instance.save()
    
    return redirect('ai_form')

# Ai call records (single call)
def call_records_single(list_data,uploder_name):
    client = ElevenLabs(api_key=settings.ELEVENLABS_API_KEY)

    conversation_id = list_data['conversation_id']
    
    
    conversation = client.conversational_ai.conversations.get(conversation_id)
    conversation_json = json.dumps(conversation.__dict__, default=str)
    conversation_dict = json.loads(conversation_json)

    print(conversation)
    analysis = conversation_dict.get("analysis", "")
    analysis1 = conversation_dict.get("conversation_initiation_client_data", "")

    # Extract called number
    match1 = re.search(r"'system__called_number':\s*'([^']+)'", analysis1)
    match2 = re.search(r"external_number='(\+\d+)'", analysis1)
    match2 = re.search(r"external_number='(\+\d+)'", conversation_dict["metadata"])
    external_number = match2.group(1) if match2 else None
    called_number = match1.group(1) if match1 else external_number

    
    # Extract transcript summary
    try:
        match = re.search(r"transcript_summary='(.*?)'", analysis)
        transcript_summary = match.group(1).encode().decode('unicode_escape')
        feedback = classify_interest(transcript_summary)
        Audio = audio(conversation_id)
        audio_filename = Audio[0]['audio_filename']
        response_content = Audio[0]['response_content']

    except Exception as e:
        print(f"Error extracting called_number: {e}")
        feedback = "**Not pick call**"
        transcript_summary = ""
        audio_filename = "NO"
        response_content = ""

    instance = call_details(
        uploader_name=uploder_name,
        # name=name,
        # address=address,
        phone_no=called_number,
        feedback=feedback,
        summary=transcript_summary,
        audio_flie=audio_filename,
        time=datetime.now()
    )

    if response_content and audio_filename.endswith('.mp3'):
        instance.audio_url.save(
            audio_filename,
            ContentFile(response_content),
            save=False
        )

    instance.save()
    
    return JsonResponse({"message": "Process completed"})


def classify_interest(summary):
    interested_keywords = [
        "interested", "wants to visit", "asked for details", "agreed to visit",
        "confirmed interest", "positive response", "requested more info", "looking forward", "interest", "curious", "inquisitive", "intrigued", "eager", "inquiring", "attentive", "investigative", "questioning",
        "nosey", "probing", "process", "engaged", "involved", "committed", "active", "absorbed", "immersed",
        "participating", "focused", "concerned", "biased", "partial", "invested", "concerned", "affected","inclined","favorable", "self-interested", "attracted", "enchanted", "fascinated", "captivated", "enthralled", "spellbound", "enthusiastic", "keen", "passionate"
    ]

    not_interested_keywords = [
        "not interested", "declined", "no interest", "not looking", "rejected",
        "not now", "maybe later", "don't want", "denied", "disinterested", "not my thing", "no thanks",
        "it's not for me", "maybe another time"
    ]

    # Check for keywords
    for phrase in not_interested_keywords:
        if phrase in summary:
            return "**Not Interested**"

    for phrase in interested_keywords:
        if phrase in summary:
            return "**Interested**"
    return "**Follow up**"

def audio(conversation_id):
    headers = {
        "xi-api-key": settings.ELEVENLABS_API_KEY,
        "Accept": "application/json"
    }
    url = f"https://api.elevenlabs.io/v1/convai/conversations/{conversation_id}/audio"
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        audio_filename = f"audios/audio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp3"
        return [{
            'audio_filename': audio_filename,
            'response_content': response.content
        }]
    else:
        print("Failed to download audio:", response.status_code)
        return [{
            'audio_filename': '',
            'response_content': b''
        }]



    
def call_history(request):
    queryset = call_details.objects.filter(phone_no__isnull=False).order_by('-id')
    data = []

    for call in queryset:
        data.append({

            "id": call.id,
            "name": call.name,
            "phone": call.phone_no,
            "date": call.date.strftime("%Y-%m-%d") if call.date else "",
            "time": call.time.strftime("%H:%M:%S") if call.time else "",
            "manager": call.uploader_name,
            "feedback": call.feedback,
            "address": call.address,
            "summary": call.summary,
            "audio_file": call.audio_flie.url if call.audio_flie and hasattr(call.audio_flie, 'url') else "",
            "audio_url": str(call.audio_url) if call.audio_url else "",
        })

    return render(request, 'call-history.html', {'call_json': json.dumps(data)})

def inboundcall(request):
    client = ElevenLabs(api_key=settings.ELEVENLABS_API_KEY)
    
    conversations_data = client.conversational_ai.conversations.list()
    
    conversation_ids = [conv.conversation_id for conv in conversations_data.conversations]
   
    data = []
    for conv in conversation_ids:
        conversation_client = client.conversational_ai.conversations.get(conv)
        conversation_json = json.dumps(conversation_client.__dict__, default=str)
        conversation_dict = json.loads(conversation_json)

        metadata = str(conversation_dict.get("metadata", ""))
        match_direction = re.search(r"direction='(\w+)'", metadata)
        direction = match_direction.group(1) if match_direction else None
        if direction == "inbound":
            print(direction)
            if not inboundcalls.objects.filter(conversation_id=conv).exists():            
                result = extract_conversation_data(conversation_dict)
                data.append(result)

    for result in data:
        instance = inboundcalls(
            agent_id=result['agent_id'],
            conversation_id=result["conversation_id"],
            summary=result["summary"],
            external_number=result["external_number"],
            direction=result["direction"],
            call_datetime=result["call_datetime"],
            
        )
        
        if result["response_content"] and result["audio_filename"].endswith('.mp3'):
            instance.audio_url.save(
                result["audio_filename"],
                ContentFile(result["response_content"]),
                save=False
            )

        instance.save()
    
    return redirect('inboundcall_loc')


def extract_conversation_data(list_data):
    client = ElevenLabs(api_key=settings.ELEVENLABS_API_KEY)
    conversation_id = list_data['conversation_id']
    conversation = client.conversational_ai.conversations.get(conversation_id)

    conversation_json = json.dumps(conversation.__dict__, default=str)
    conversation_dict = json.loads(conversation_json)

    agent_id = conversation_dict.get("agent_id")
    status = conversation_dict.get("status")
    call_duration_secs = conversation_dict.get("call_duration_secs")

    analysis = str(conversation_dict.get("analysis", {}))
    match_summary = re.search(r"transcript_summary='(.*?)'", analysis)
    
    summary = match_summary.group(1) if match_summary else ""
    audio_filename =""
    response_content=""
    # if summary:
    Audio = audio(conversation_id)
    audio_filename = Audio[0]['audio_filename']
    response_content = Audio[0]['response_content']

    
    metadata = str(conversation_dict.get("metadata", ""))
    metadata_str = str(metadata)
    match_number = re.search(r"external_number='(\+\d+)'", metadata_str)
    external_number = match_number.group(1) if match_number else None

    match_direction = re.search(r"direction='(\w+)'", metadata_str)
    direction = match_direction.group(1) if match_direction else None

    match_time = re.search(r"start_time_unix_secs=(\d+)", str(metadata))
    
    if match_time:
        start_time_unix = int(match_time.group(1))
        call_datetime = datetime.utcfromtimestamp(int(start_time_unix)).strftime('%Y-%m-%d %H:%M:%S')
        print(call_datetime)
    else:
        call_datetime = None
        

    match_audio = re.search(r"recording_url=['\"]([^'\"]+)['\"]", metadata_str)
    audio_url = match_audio.group(1) if match_audio else None
    print("Audio URL:", audio_url)
    # pprint.pprint(conversation_dict.get("metadata"))

    init_data = conversation_dict.get("conversation_initiation_client_data", "")
    match_called = re.search(r"'system__called_number':\s*'([^']+)'", str(init_data))
    called_number = match_called.group(1) if match_called else external_number


    # Final formatted output
    result = {
        "agent_id": agent_id,
        "conversation_id": conversation_id,
        "summary": summary,
        "external_number": external_number,
        "direction": direction,
        "call_datetime":call_datetime,
        "audio_url":audio_url,
        "audio_filename":audio_filename,
        "response_content":response_content
    }
    return result
def inboundcall_loc(request):
    queryset = inboundcalls.objects.all().order_by('-id')
    data = []

    for call in queryset:
        data.append({
            "id": call.id,
            "summary": call.summary,
            "external_number": call.external_number,
            "direction": call.direction,
            "audio_url": call.audio_url.url if call.audio_url else '',
            "date": str(call.call_datetime) if call.call_datetime else '',
        })


    return render(request, 'inbound_call.html', {'call_json': json.dumps(data)})

def widget_calls_loc(request):
    queryset = WidgetCalls.objects.all().order_by("-id")

    data = []
    for call in queryset:
        data.append({
            "id": call.id,

            # UI expects YYYY-MM-DD
            "date": call.call_datetime.strftime("%Y-%m-%d")
                    if call.call_datetime else "",

            # Widget calls have NO phone number
            # So we fake one for UI consistency
            "external_number": "Widget Call",

            # UI filter expects direction
            "direction": "Inbound",

            # UI audio player expects direct URL
            "audio_url": (
                call.audio_file.url
                if call.audio_file
                else ""
            ),
        })

    return render(
        request,
        "widget_call.html",   # your same HTML
        {"call_json": json.dumps(data)}
    )

def widget_calls(request):
    client = ElevenLabs(api_key=os.getenv("ELEVENLABS_API_KEY"))

    conversations = client.conversational_ai.conversations.list()

    data = []

    for conv in conversations.conversations:
        conv_id = conv.conversation_id

        # Skip already saved
        if WidgetCalls.objects.filter(conversation_id=conv_id).exists():
            continue

        conversation = client.conversational_ai.conversations.get(conv_id)
        conv_dict = json.loads(json.dumps(conversation.__dict__, default=str))

        metadata = str(conv_dict.get("metadata", ""))

        # ✅ Widget calls do NOT have external_number
        match_number = re.search(r"external_number='(\+\d+)'", metadata)
        external_number = match_number.group(1) if match_number else None

        if external_number:
            continue  # ❌ skip phone calls

        result = extract_widget_data(conv_dict)

        if result:
            data.append(result)

    for item in data:
        instance = WidgetCalls(
            agent_id=item["agent_id"],
            conversation_id=item["conversation_id"],
            summary=item["summary"],
            call_datetime=item["call_datetime"],
            audio_url=item["audio_url"],
        )

        if item["audio_content"]:
            instance.audio_file.save(
                item["audio_filename"],
                ContentFile(item["audio_content"]),
                save=False
            )

        instance.save()

    return redirect("widget_calls")

def extract_widget_data(conversation_dict):
    conversation_id = conversation_dict.get("conversation_id")
    agent_id = conversation_dict.get("agent_id")

    # -------- SUMMARY ----------
    analysis = str(conversation_dict.get("analysis", ""))
    match_summary = re.search(r"transcript_summary='(.*?)'", analysis)
    summary = match_summary.group(1) if match_summary else ""

    # -------- METADATA ----------
    metadata = str(conversation_dict.get("metadata", ""))

    # Recording URL (THIS IS WHAT YOU NEED)
    match_audio = re.search(r"recording_url=['\"]([^'\"]+)['\"]", metadata)
    recording_url = match_audio.group(1) if match_audio else None

    if not recording_url:
        return None

    # -------- TIME ----------
    match_time = re.search(r"start_time_unix_secs=(\d+)", metadata)
    call_datetime = None
    if match_time:
        call_datetime = datetime.utcfromtimestamp(
            int(match_time.group(1))
        )

    # -------- DOWNLOAD AUDIO ----------
    audio_content = None
    audio_filename = f"{conversation_id}.mp3"

    try:
        response = requests.get(recording_url, timeout=15)
        if response.status_code == 200:
            audio_content = response.content
    except Exception as e:
        print("Audio download error:", e)

    return {
        "agent_id": agent_id,
        "conversation_id": conversation_id,
        "summary": summary,
        "call_datetime": call_datetime,
        "audio_url": recording_url,
        "audio_filename": audio_filename,
        "audio_content": audio_content,
    }
