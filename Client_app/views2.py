from django.shortcuts import render
from django.shortcuts import render
from openpyxl import load_workbook
from openpyxl import Workbook
import cv2
import numpy as np
from django.conf import settings
from django.http import JsonResponse
from twilio.rest import Client
from nltk.sentiment import SentimentIntensityAnalyzer
import os
import re
import json
import pandas as pd
import httpx
from django.http import HttpResponse
import time
from .models import *
from datetime import datetime
from pathlib import Path
import requests
from io import BytesIO
import math
from django.core.files.base import ContentFile

from elevenlabs import ElevenLabs, OutboundCallRecipient
from django.utils.timezone import now
from django.shortcuts import redirect
from dotenv import load_dotenv
from .hubspot_integration import sync_to_hubspot, HubSpotCRMIntegration
import logging
from django.views.decorators.csrf import csrf_exempt
from django.http import StreamingHttpResponse, HttpResponse
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Prefetch
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.db.models import Count



load_dotenv()

logger = logging.getLogger(__name__)


def home(request):
    today = now().date()
    print(today)
    if request.user.is_authenticated and request.user.role == "admin":
        Total_count = call_details.objects.filter(phone_no__isnull=False, user_data=request.user.id).count()
        # print(Total_count)
        inbound_count = inboundcalls.objects.filter(user_data=request.user.id).count()
        widget_data = ElevenCall.objects.filter(user_data=request.user.id).count()
        total_count = Total_count + inbound_count + widget_data
        # print(Interested_count)
        Not_pick_count = call_details.objects.filter(feedback="**Not pick call**", phone_no__isnull=False).count()
        today_count = call_details.objects.filter(date = today, phone_no__isnull=False, user_data=request.user.id).count()
        data = call_details.objects.filter(feedback="**Interested**", user_data=request.user.id).order_by('-date')[:5]
        return render(request, "index.html", {'data':data, 'Total_count':Total_count, 'inbound_count':inbound_count, 'Not_pick_count':Not_pick_count, 'today_count':today_count, 'name':'Mathavan', 'widget_data':widget_data, 'total_count':total_count})

    Total_count = call_details.objects.filter(phone_no__isnull=False).count()
    # print(Total_count)
    inbound_count = inboundcalls.objects.all().count()
    widget_data = ElevenCall.objects.all().count()
    total_count = Total_count + inbound_count + widget_data
    # print(Interested_count)
    Not_pick_count = call_details.objects.filter(feedback="**Not pick call**", phone_no__isnull=False).count()
    today_count = call_details.objects.filter(date = today, phone_no__isnull=False).count()
    data = call_details.objects.filter(feedback="**Interested**").order_by('-date')[:5]
    return render(request, "index.html", {'data':data, 'Total_count':Total_count, 'inbound_count':inbound_count, 'Not_pick_count':Not_pick_count, 'today_count':today_count, 'name':'Mathavan', 'widget_data':widget_data, 'total_count':total_count})

def ai_form(requests):
    return render(requests, "ai-calling.html")

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("home")
        return render(request, "auth-login.html",{'message':"invalid username or password"})
    return render(request, "auth-login.html")

def logout_view(request):
    logout(request)
    return render(request, "auth-login.html")

def call_process(request):
    if request.method == 'POST':
        Name = request.POST['name']
        Time = request.POST['time']
        Date = request.POST['date']
        try:
            excel_file = request.FILES['file']
            wb = load_workbook(excel_file)
            ws = wb.active
            call_list =[]
            for row in ws.iter_rows(min_row=2, values_only=True):
                print(row[0], row[1], row[2], row[3])
                ai_call = generate_speech(row[0], row[1], row[2]) 
            
                call_list.append(ai_call)
            time.sleep(80)
            call_records(call_list, Name)
            return render(request, 'ai-calling.html')
        except:
            excel_file = request.POST['number']
            ai_call =  generate_speech_single(excel_file)
        
            time.sleep(120)
            call_records_single(ai_call, Name)
            return render(request, 'ai-calling.html')
    
    return render(request, "ai-calling.html")

def generate_speech(name, address, to_number):
    try:
        client = ElevenLabs(api_key=os.getenv('ELEVENLABS_API_KEY'))
        phone_numbers = client.conversational_ai.phone_numbers.list()

        for number in phone_numbers:
            print(vars(number))

        response = client.conversational_ai.twilio.outbound_call(
                    agent_id=settings.ELEVENLABS_AGENT_DUBAI,
                    agent_phone_number_id=os.getenv('AGENT_PHONE_NUMBER_ID_DUBAI'),
                    to_number= f"+91{to_number}",
                )
        
        
        data ={'conversation_id':response.conversation_id,
               'name':name,
               'address':address,
               }
        return data

    except Exception as e:
        return JsonResponse({"error": "Error while scheduling call", "details": str(e)}, status=500)


def generate_speech_single(to_number):
    try:
        # to_number = 9600388948
        # to_number = 9600388948
        client = ElevenLabs(api_key=os.getenv('ELEVENLABS_API_KEY'))
        phone_numbers = client.conversational_ai.phone_numbers.list()

        for number in phone_numbers:
            print(vars(number))

        response = client.conversational_ai.twilio.outbound_call(
                    agent_id=settings.ELEVENLABS_AGENT_DUBAI,
                    agent_phone_number_id=os.getenv('AGENT_PHONE_NUMBER_ID_DUBAI'),
                    to_number= f"+91{to_number}",
                )
        
        
        data ={'conversation_id':response.conversation_id
               }
        return data
        # return redirect('home')
        

    except Exception as e:
        return JsonResponse({"error": "Error while scheduling call", "details": str(e)}, status=500)


def call_records(data, uploder_name, request):

    client = ElevenLabs(api_key=os.getenv('ELEVENLABS_API_KEY'))

    all_results = []
    for list_data in data:
        conversation_id = list_data['conversation_id']
        name = list_data['name']
        address = list_data['address']
       
        conversation = client.conversational_ai.conversations.get(conversation_id)
        conversation_json = json.dumps(conversation.__dict__, default=str)
        conversation_dict = json.loads(conversation_json)

        print(conversation)
        analysis = conversation_dict.get("analysis", "")
        analysis1 = conversation_dict.get("conversation_initiation_client_data", "")

       
        match1 = re.search(r"'system__called_number':\s*'([^']+)'", analysis1)
        match2 = re.search(r"external_number='(\+\d+)'", analysis1)
        match2 = re.search(r"external_number='(\+\d+)'", conversation_dict["metadata"])
        external_number = match2.group(1) if match2 else None
        called_number = match1.group(1) if match1 else external_number

        
       
        try:
            match = re.search(r"transcript_summary='(.*?)'", analysis)
            transcript_summary = match.group(1).encode().decode('unicode_escape')
            feedback = classify_interest(transcript_summary)
            Audio = audio(conversation_id)
            audio_filename = Audio[0]['audio_filename']
            response_content = Audio[0]['response_content']

        except Exception as e:
            print(f"Error extracting called_number: {e}")
            feedback = "**Not pick call**"
            transcript_summary = ""
            audio_filename = "NO"
            # audio_filename = ""
            response_content =""

       
        all_results.append({
            "Name": name,
            "Address": address,
            "Called Number": called_number,
            "Feedback": feedback,
            "Transcript Summary": transcript_summary,
            "audio_filename":audio_filename,
            "response_content":response_content
            
        })
    for result in all_results:
        instance = call_details(
            uploader_name=uploder_name,
            name=result["Name"],
            address=result["Address"],
            phone_no=result["Called Number"],
            feedback=result["Feedback"],
            summary=result["Transcript Summary"],
            audio_flie=result["audio_filename"],
            time=datetime.now(),
            user_data=request.user.id
        )
        
        if result["response_content"] and result["audio_filename"].endswith('.mp3'):
            instance.audio_url.save(
                result["audio_filename"],
                ContentFile(result["response_content"]),
                save=False
            )

        instance.save()
        
        
        try:
            sync_success = sync_to_hubspot(instance)
            if sync_success:
                logger.info(f"Successfully synced call details to HubSpot for {instance.phone_no}")
            else:
                logger.warning(f"Failed to sync call details to HubSpot for {instance.phone_no}")
        except Exception as e:
            logger.error(f"Error syncing to HubSpot: {str(e)}")
    
    return redirect('ai_form')

def call_records_single(list_data,uploder_name):
    client = ElevenLabs(api_key=os.getenv('ELEVENLABS_API_KEY'))

    conversation_id = list_data['conversation_id']
    
    
    conversation = client.conversational_ai.conversations.get(conversation_id)
    conversation_json = json.dumps(conversation.__dict__, default=str)
    conversation_dict = json.loads(conversation_json)

    print(conversation)
    analysis = conversation_dict.get("analysis", "")
    analysis1 = conversation_dict.get("conversation_initiation_client_data", "")


    match1 = re.search(r"'system__called_number':\s*'([^']+)'", analysis1)
    match2 = re.search(r"external_number='(\+\d+)'", analysis1)
    match2 = re.search(r"external_number='(\+\d+)'", conversation_dict["metadata"])
    external_number = match2.group(1) if match2 else None
    called_number = match1.group(1) if match1 else external_number


    try:
        match = re.search(r"transcript_summary='(.*?)'", analysis)
        transcript_summary = match.group(1).encode().decode('unicode_escape')
        feedback = classify_interest(transcript_summary)
        Audio = audio(conversation_id)
        audio_filename = Audio[0]['audio_filename']
        response_content = Audio[0]['response_content']

    except Exception as e:
        print(f"Error extracting called_number: {e}")
        feedback = "**Not pick call**"
        transcript_summary = ""
        audio_filename = "NO"
        response_content = ""

    instance = call_details(
        uploader_name=uploder_name,
        # name=name,
        # address=address,
        phone_no=called_number,
        feedback=feedback,
        summary=transcript_summary,
        audio_flie=audio_filename,
        time=datetime.now()
    )

    if response_content and audio_filename.endswith('.mp3'):
        instance.audio_url.save(
            audio_filename,
            ContentFile(response_content),
            save=False
        )

    instance.save()
    
    
    try:
        sync_success = sync_to_hubspot(instance)
        if sync_success:
            logger.info(f"Successfully synced call details to HubSpot for {instance.phone_no}")
        else:
            logger.warning(f"Failed to sync call details to HubSpot for {instance.phone_no}")
    except Exception as e:
        logger.error(f"Error syncing to HubSpot: {str(e)}")
    
    return JsonResponse({"message": "Process completed"})


def classify_interest(summary):
    interested_keywords = [
        "interested", "wants to visit", "asked for details", "agreed to visit",
        "confirmed interest", "positive response", "requested more info", "looking forward", "interest", "curious", "inquisitive", "intrigued", "eager", "inquiring", "attentive", "investigative", "questioning",
        "nosey", "probing", "process", "engaged", "involved", "committed", "active", "absorbed", "immersed",
        "participating", "focused", "concerned", "biased", "partial", "invested", "concerned", "affected","inclined","favorable", "self-interested", "attracted", "enchanted", "fascinated", "captivated", "enthralled", "spellbound", "enthusiastic", "keen", "passionate"
    ]

    not_interested_keywords = [
        "not interested", "declined", "no interest", "not looking", "rejected",
        "not now", "maybe later", "don't want", "denied", "disinterested", "not my thing", "no thanks",
        "it's not for me", "maybe another time"
    ]

    
    for phrase in not_interested_keywords:
        if phrase in summary:
            return "**Not Interested**"

    for phrase in interested_keywords:
        if phrase in summary:
            return "**Interested**"
    return "**Follow up**"

def audio(conversation_id):
    headers = {
        "xi-api-key": os.getenv('ELEVENLABS_API_KEY'),
        "Accept": "application/json"
    }
    url = f"https://api.elevenlabs.io/v1/convai/conversations/{conversation_id}/audio"
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        audio_filename = f"audios/audio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp3"
        return [{
            'audio_filename': audio_filename,
            'response_content': response.content
        }]
    else:
        print("Failed to download audio:", response.status_code)
        return [{
            'audio_filename': '',
            'response_content': b''
        }]



    
from django.utils.timezone import now
import json

def call_history(request):
    filter_type = request.GET.get('type')  # 👈 get filter from URL
    if request.user.is_authenticated and request.user.role == "admin":
        queryset = call_details.objects.filter(phone_no__isnull=False, user_data=request.user.id)

        # ===== FILTER LOGIC =====
        if filter_type == "**Interested**":
            queryset = queryset.filter(feedback__iexact="**Interested**")
            print(queryset)

        elif filter_type == "**Not Interested**":
            queryset = queryset.filter(feedback__iexact="**Not Interested**")

        elif filter_type == "today":
            today = now().date()
            queryset = queryset.filter(date=today)

        # else → total calls (no filter)

        queryset = queryset.order_by('-id')

        data = []
        for call in queryset:
            data.append({
                "id": call.id,
                "name": call.name,
                "phone": call.phone_no,
                "date": call.date.strftime("%Y-%m-%d") if call.date else "",
                "time": call.time.strftime("%H:%M:%S") if call.time else "",
                "manager": call.uploader_name,
                "feedback": call.feedback,
                "address": call.address,
                "summary": call.summary,
                "audio_file": call.audio_flie.url if call.audio_flie and hasattr(call.audio_flie, 'url') else "",
                "audio_url": str(call.audio_url) if call.audio_url else "",
            })


        return render(request, 'call-history.html', {'call_json': json.dumps(data)})
    queryset = call_details.objects.filter(phone_no__isnull=False)

    # ===== FILTER LOGIC =====
    if filter_type == "**Interested**":
        queryset = queryset.filter(feedback__iexact="**Interested**")
        print(queryset)

    elif filter_type == "**Not Interested**":
        queryset = queryset.filter(feedback__iexact="**Not Interested**")

    elif filter_type == "today":
        today = now().date()
        queryset = queryset.filter(date=today)

    # else → total calls (no filter)

    queryset = queryset.order_by('-id')

    data = []
    for call in queryset:
        data.append({
            "id": call.id,
            "name": call.name,
            "phone": call.phone_no,
            "date": call.date.strftime("%Y-%m-%d") if call.date else "",
            "time": call.time.strftime("%H:%M:%S") if call.time else "",
            "manager": call.uploader_name,
            "feedback": call.feedback,
            "address": call.address,
            "summary": call.summary,
            "audio_file": call.audio_flie.url if call.audio_flie and hasattr(call.audio_flie, 'url') else "",
            "audio_url": str(call.audio_url) if call.audio_url else "",
        })


    return render(request, 'call-history.html', {'call_json': json.dumps(data)})

import requests

def download_audio(url):
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return response.content

def inboundcall(request):
    client = ElevenLabs(api_key=settings.ELEVENLABS_API_KEY)

    conversations = client.conversational_ai.conversations.list(agent_id=settings.ELEVENLABS_AGENT_DUBAI).conversations

    print(conversations)
    for conv in conversations:
        conversation_id = conv.conversation_id

        # Prevent duplicate insert
        if inboundcalls.objects.filter(conversation_id=conversation_id).exists():
            continue

        conversation = client.conversational_ai.conversations.get(conversation_id)
        result = extract_conversation_data(conversation)
        print("DEBUG direction:", result["direction"])
        # Store only inbound calls
        if result["direction"] != "inbound":
            continue

        instance = inboundcalls(
            agent_id=result["agent_id"],
            conversation_id=result["conversation_id"],
            summary=result["summary"],
            external_number=result["external_number"],
            direction=result["direction"],
            call_datetime=result["call_datetime"],
            user_data=request.user.id
        )

        # Save audio file
        if result["audio_content"]:
            instance.audio_url.save(
                result["audio_filename"],
                ContentFile(result["audio_content"]),
                save=False
            )

        instance.save()

    return redirect("inboundcall_loc")



def extract_conversation_data(conversation):
    metadata = str(conversation.metadata or "")
    analysis = str(conversation.analysis or "")

    # Direction
    print(conversation)
    match_direction = re.search(r"direction='(\w+)'", metadata)
    direction = match_direction.group(1) if match_direction else ""

    # External number
    match_number = re.search(r"external_number='(\+\d+)'", metadata)
    external_number = match_number.group(1) if match_number else ""

    # Call datetime
    match_time = re.search(r"start_time_unix_secs=(\d+)", metadata)
    if match_time:
        call_datetime = datetime.utcfromtimestamp(
            int(match_time.group(1))
        ).strftime('%Y-%m-%d %H:%M:%S')
    else:
        call_datetime = ""

    # Summary
    match_summary = re.search(r"transcript_summary='(.*?)'", analysis)
    summary = match_summary.group(1) if match_summary else ""

    # Audio
    audio_filename = ""
    audio_content = None

    match_audio = re.search(r"recording_url=['\"]([^'\"]+)['\"]", metadata)
    if match_audio:
        audio_url = match_audio.group(1)
        print("audio_url")
        audio_filename = f"{conversation.conversation_id}.mp3"
        print("audio_filename :", audio_filename)
        audio_content = download_audio(audio_url)

    return {
        "agent_id": conversation.agent_id,
        "conversation_id": conversation.conversation_id,
        "summary": summary,
        "external_number": external_number,
        "direction": direction,
        "call_datetime": call_datetime,
        "audio_filename": audio_filename,
        "audio_content": audio_content,
    }

def inboundcall_loc(request):
    if request.user.is_authenticated and request.user.role == "admin":
        queryset = inboundcalls.objects.filter(user_data=request.user.id).order_by('-id')
        data = []

        for call in queryset:
            data.append({
                "id": call.id,
                "summary": call.summary,
                "external_number": call.external_number,
                "direction": call.direction,
                "summary": call.summary,
                "audio_url": call.audio_url.url if call.audio_url else '',
                "date": str(call.call_datetime) if call.call_datetime else '',
            })


        return render(request, 'inbound_call.html', {'call_json': json.dumps(data)})
    queryset = inboundcalls.objects.all().order_by('-id')
    data = []

    for call in queryset:
        data.append({
            "id": call.id,
            "summary": call.summary,
            "external_number": call.external_number,
            "direction": call.direction,
            "summary": call.summary,
            "audio_url": call.audio_url.url if call.audio_url else '',
            "date": str(call.call_datetime) if call.call_datetime else '',
        })


    return render(request, 'inbound_call.html', {'call_json': json.dumps(data)})


# def call_process(request):
#     if request.method == 'POST':
#         Name = request.POST['name']
#         Time = request.POST['time']
#         Date = request.POST['date']
#         try:
#             excel_file = request.FILES['file']
#             wb = load_workbook(excel_file)
#             ws = wb.active
#             call_list =[]
#             for row in ws.iter_rows(min_row=2, values_only=True):
#                 print(row[0], row[1], row[2], row[3])
#                 ai_call = generate_speech(row[0], row[1], row[2]) 
            
#                 call_list.append(ai_call)
#             time.sleep(80)
#             call_records(call_list, Name)
#             return render(request, 'ai-calling.html')
#         except:
#             excel_file = request.POST['number']
#             ai_call =  generate_speech_single(excel_file)
        
#             time.sleep(100)
#             call_records_single(ai_call, Name)
#             return render(request, 'ai-calling.html')
    
#     return render(request, "ai-calling.html")

def hubspot_config(request):
    config = HubSpotConfig.objects.filter(is_active=True).first()
    return render(request, 'hubspot_config.html', {'config': config})

@csrf_exempt
def save_hubspot_config(request):
    if request.method == 'POST':
        try:
            access_token = request.POST.get('access_token')
            portal_id = request.POST.get('portal_id')
            
            if not access_token or not portal_id:
                return JsonResponse({'success': False, 'message': 'Access token and portal ID are required'})
            
            HubSpotConfig.objects.filter(is_active=True).update(is_active=False)
            
            config = HubSpotConfig.objects.create(
                access_token=access_token,
                portal_id=portal_id,
                is_active=True
            )
            
            hubspot = HubSpotCRMIntegration()
            
            test_data = {
                'name': 'Test Contact',
                'phone_no': '+1234567890',
                'address': 'Test Address',
                'feedback': 'Interested',
                'summary': 'Test sync from configuration',
                'date': datetime.now().date(),
                'time': datetime.now().time(),
                'uploader_name': 'System Test'
            }
            
            class TestCallDetail:
                def __init__(self, data):
                    for key, value in data.items():
                        setattr(self, key, value)
            
            test_call = TestCallDetail(test_data)
            
            try:
                response = requests.get(f"https://api.hubapi.com/crm/v3/objects/contacts", 
                                      headers=hubspot.headers, 
                                      params={'limit': 1})
                if response.status_code == 200:
                    return JsonResponse({
                        'success': True, 
                        'message': 'HubSpot configuration saved and tested successfully!'
                    })
                else:
                    config.delete()
                    return JsonResponse({
                        'success': False, 
                        'message': f'Invalid credentials. Status code: {response.status_code}'
                    })
            except Exception as e:
                config.delete()
                return JsonResponse({
                    'success': False, 
                    'message': f'Failed to connect to HubSpot: {str(e)}'
                })
                
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error saving configuration: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

def sync_existing_data(request):
    if request.method == 'POST':
        try:
            config = HubSpotConfig.objects.filter(is_active=True).first()
            if not config:
                return JsonResponse({'success': False, 'message': 'No active HubSpot configuration found'})
            
            call_details_list = call_details.objects.filter(phone_no__isnull=False)[:10]
            
            success_count = 0
            for call_detail in call_details_list:
                try:
                    if sync_to_hubspot(call_detail):
                        success_count += 1
                except Exception as e:
                    logger.error(f"Failed to sync call detail {call_detail.id}: {str(e)}")
            
            return JsonResponse({
                'success': True, 
                'message': f'Successfully synced {success_count} out of {len(call_details_list)} records to HubSpot'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error syncing data: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@csrf_exempt
def sync_hubspot_leads(request):
    """Sync leads from HubSpot to local database"""
    if request.method == 'POST':
        try:
            from .hubspot_integration import HubSpotCRMIntegration
            
            config = HubSpotConfig.objects.filter(is_active=True).first()
            if not config:
                return JsonResponse({'success': False, 'message': 'No active HubSpot configuration found'})
            
            hubspot = HubSpotCRMIntegration()
            synced_count = hubspot.sync_contacts_from_hubspot(incremental=True)
            
            return JsonResponse({
                'success': True, 
                'message': f'Successfully synced {synced_count} new leads from HubSpot'
            })
            
        except Exception as e:
            logger.error(f"Error syncing leads from HubSpot: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Error syncing leads: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@csrf_exempt  
def auto_sync_hubspot(request):
    """Automatic sync endpoint for periodic calls (can be called by cron jobs or webhooks)"""
    try:
        from .hubspot_integration import HubSpotCRMIntegration
        
        config = HubSpotConfig.objects.filter(is_active=True).first()
        if not config:
            return JsonResponse({'success': False, 'message': 'No active HubSpot configuration found'})
        
        hubspot = HubSpotCRMIntegration()
        synced_count = hubspot.sync_contacts_from_hubspot(incremental=True)
        
        logger.info(f"Auto-sync completed: {synced_count} leads synced from HubSpot")
        
        return JsonResponse({
            'success': True, 
            'synced_count': synced_count,
            'message': f'Auto-sync completed: {synced_count} leads synced'
        })
        
    except Exception as e:
        logger.error(f"Error in auto-sync: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Auto-sync failed: {str(e)}'})


import json
from datetime import date
from django.utils import timezone


def widget_calls_loc(request):
    if request.user.is_authenticated and request.user.role == "admin":
        queryset = ElevenCall.objects.filter(user_data=request.user.id).order_by("-call_date")

        data = []
        for call in queryset:
            data.append({
                "id": call.id,
                "date": call.call_date.strftime("%Y-%m-%d")
                        if call.call_date else "",
                "external_number": "Widget Call",
                "time": call.call_time.strftime("%H:%M:%S")
                if call.call_time else "",
                "direction": call.direction or "Widget",
                "conversation_id": call.conversation_id,
                "summary": call.summary or "",
                "has_audio": call.has_audio,
                "transcript":call.transcript
            })

        return render(
            request,
            "widget.html",
            {
                "call_json": json.dumps(data)
            }
        )



    queryset = ElevenCall.objects.all().order_by("-call_date")

    data = []
    for call in queryset:
        data.append({
            "id": call.id,
            "date": call.call_date.strftime("%Y-%m-%d")
                    if call.call_date else "",
            "external_number": "Widget Call",
            "time": call.call_time.strftime("%H:%M:%S")
            if call.call_time else "",
            "direction": call.direction or "Widget",
            "conversation_id": call.conversation_id,
            "summary": call.summary or "",
            "has_audio": call.has_audio,
            "transcript":call.transcript
        })

    return render(
        request,
        "widget.html",
        {
            "call_json": json.dumps(data)
        }
    )

def widget_calls_api(request):

    if not request.user.is_authenticated:
        return JsonResponse({"error": "Unauthorized"}, status=401)

    search = request.GET.get("search", "")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    page = request.GET.get("page", 1)

    try:
        page = int(page)
    except:
        page = 1

    queryset = ElevenCall.objects.all().order_by("-call_date")

    # Admin restriction (your logic)
    if request.user.role == "admin":
        queryset = queryset.filter(user_data=request.user.id)

    # 🔎 Search filter
    if search:
        queryset = queryset.filter(
            Q(direction__icontains=search)
        )

    # 📅 Date Range Filter
    if from_date and to_date:
        queryset = queryset.filter(call_date__range=[from_date, to_date])
    elif from_date:
        queryset = queryset.filter(call_date__gte=from_date)
    elif to_date:
        queryset = queryset.filter(call_date__lte=to_date)

    paginator = Paginator(queryset, 10)

    try:
        page_obj = paginator.page(page)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    data = []
    for call in page_obj:
        data.append({
            "id": call.id,
            "date": call.call_date.strftime("%Y-%m-%d") if call.call_date else "",
            "time": call.call_time.strftime("%H:%M:%S") if call.call_time else "",
            "direction": call.direction or "Widget",
            "conversation_id": call.conversation_id,
            "external_number": "Widget Call",
            "has_audio": call.has_audio,
            "transcript": call.transcript or ""
        })

    return JsonResponse({
        "data": data,
        "total_pages": paginator.num_pages,
        "current_page": page_obj.number
    })
# =========================
# MAIN VIEW
# =========================

def download_call_audio(conversation_id):
    url = f"https://api.elevenlabs.io/v1/convai/conversations/{conversation_id}/audio"
    headers = {"xi-api-key": settings.ELEVENLABS_API_KEY}

    r = requests.get(url, headers=headers)
    if r.status_code != 200:
        return None

    folder = os.path.join(settings.MEDIA_ROOT, "call_audio")
    os.makedirs(folder, exist_ok=True)

    file_path = os.path.join(folder, f"{conversation_id}.mp3")
    with open(file_path, "wb") as f:
        f.write(r.content)

    return f"call_audio/{conversation_id}.mp3"

def fetch_all_conversations(client, agent_id):
    all_conversations = []
    cursor = None

    while True:
        response = client.conversational_ai.conversations.list(
            agent_id=agent_id,
            cursor=cursor
        )

        all_conversations.extend(response.conversations)

        if not response.next_cursor:
            break

        cursor = response.next_cursor

    return all_conversations

def widget_calls(request):
    client = ElevenLabs(api_key=settings.ELEVENLABS_API_KEY)

    conversations = fetch_all_conversations(
        client,
        settings.ELEVENLABS_AGENT_HEALTHCARE
    )
    START_DATE = timezone.now().date()   # current date
    END_DATE   = date(2026, 1, 29)
    print("START_DATE :", START_DATE)
    print("END_DATE :", END_DATE)

    for conv in conversations:
        if ElevenCall.objects.filter(conversation_id=conv.conversation_id).exists():
            continue

        full_conv = client.conversational_ai.conversations.get(conv.conversation_id)
        user_id=request.user.id
        data = extract_widget_data(full_conv, user_id)
        call_date = data["call_date"]
        print(call_date)

        
        if not call_date:
            continue

       
        if call_date > START_DATE:
            continue

        if call_date < END_DATE:
            continue

        print("okay")
        audio_file_path = None
        if getattr(full_conv, "has_audio", False):
            audio_file_path = download_call_audio(conv.conversation_id)

        data = extract_widget_data(full_conv, user_id)

        ElevenCall.objects.create(
            conversation_id=data["conversation_id"],
            agent_id=data["agent_id"],
            direction=data["direction"],
            call_date=data["call_date"],
            call_time=data["call_time"],
            summary=data.get("summary"),
            transcript=data.get("transcript", ""),
            audio_file=audio_file_path,
            has_audio=bool(audio_file_path),
            raw_data=data["raw_data"],
            user_data=data["user_id"]
        )

    queryset = ElevenCall.objects.all().order_by("-call_date")

    data = []
    for call in queryset:
        data.append({
            "id": call.id,
            "date": call.call_date.strftime("%Y-%m-%d") if call.call_date else "",
            "time": call.call_time.strftime("%H:%M:%S") if call.call_time else "",
            "external_number": "Widget Call",
            "direction": call.direction or "Widget",
            "conversation_id": call.conversation_id,
            "summary": call.summary or "",
            "has_audio": call.has_audio,
            "transcript": call.transcript,
        })

    return render(
        request,
        "widget.html",
        {"call_json": json.dumps(data)}
    )


# =========================
# AUDIO URL EXTRACTOR
# =========================
def extract_audio_url(conversation):
    # Direct fields (if present)
    if hasattr(conversation, "audio_url") and conversation.audio_url:
        return conversation.audio_url

    if hasattr(conversation, "recording_url") and conversation.recording_url:
        return conversation.recording_url

    # Metadata is a Pydantic model → convert to dict
    if hasattr(conversation, "metadata") and conversation.metadata:
        metadata_dict = conversation.metadata.model_dump()
        return metadata_dict.get("audio_url")

    return None



# =========================
# TRANSCRIPT EXTRACTOR
# =========================
def extract_transcript(conversation):
    transcript_text = ""

    if hasattr(conversation, "transcript") and conversation.transcript:
        lines = []
        for item in conversation.transcript:
            role = getattr(item, "role", "unknown")
            message = getattr(item, "message", "")
            lines.append(f"{role.upper()}: {message}")

        transcript_text = "\n".join(lines)

    return transcript_text


# =========================
# MAIN DATA EXTRACTOR
# =========================

import re
from datetime import datetime
from django.utils.timezone import make_aware

def extract_call_date_time_from_metadata(conversation):
    metadata = str(conversation.metadata or "")

    match_time = re.search(r"start_time_unix_secs=(\d+)", metadata)

    if match_time:
        dt = datetime.utcfromtimestamp(int(match_time.group(1)))
        dt = make_aware(dt)   # timezone-aware

        call_date = dt.date()   # 👉 Date only
        call_time = dt.time()   # 👉 Time only

        return call_date, call_time

    return None, None

def extract_summary(conversation):
    if not hasattr(conversation, "analysis") or not conversation.analysis:
        return None

    analysis_dict = conversation.analysis.model_dump()

    return analysis_dict.get("summary")

def extract_widget_data(conversation, user_id):
    call_date, call_time = extract_call_date_time_from_metadata(conversation)
    return {
        "conversation_id": conversation.conversation_id,
        "agent_id": conversation.agent_id,
        "direction": (
            conversation.phone_call.direction
            if hasattr(conversation, "phone_call") and conversation.phone_call
            else None
        ),
        "call_date": call_date,
        "call_time": call_time,
        "summary": extract_summary(conversation),
        "transcript": extract_transcript(conversation),  # ✅ ALWAYS EXISTS
        "audio_url": extract_audio_url(conversation),
        "raw_data": conversation.model_dump(),
        "user_id":user_id
    }

def stream_call_audio(request, conversation_id):
    url = f"https://api.elevenlabs.io/v1/convai/conversations/{conversation_id}/audio"

    headers = {
        "xi-api-key": settings.ELEVENLABS_API_KEY
    }

    r = requests.get(url, headers=headers, stream=True)

    if r.status_code != 200:
        return HttpResponse("Audio not available", status=404)

    response = StreamingHttpResponse(
        r.iter_content(chunk_size=8192),
        content_type="audio/mpeg"
    )
    response["Content-Disposition"] = "inline; filename=call_audio.mp3"
    return response
    
def client(request):

    if request.user.is_authenticated and request.user.role == "admin":
        tenants = Tenant.objects.filter(owner=request.user)
        template_name = "admin_tenant.html"
    else:
        tenants = Tenant.objects.all()
        template_name = "clients.html"

    # ✅ Annotate branch count
    tenants = tenants.prefetch_related('branches').annotate(
        branch_count=Count('branches')
    )

    data = []

    for tenant in tenants:
        branch_list = [
            {
                "branch_name": branch.name,
                "branch_code": branch.branch_code
            }
            for branch in tenant.branches.all()
        ]

        data.append({
            "id": tenant.id,
            "tenant_name": tenant.tenant_name,
            "tenant_email": tenant.email,
            "tenant_phone": tenant.phone,
            "branch_count": tenant.branch_count,   # ✅ count
            "branches": branch_list,               # ✅ names + codes
            "date": tenant.created_at.strftime("%Y-%m-%d") if tenant.created_at else None,
        })

    context = {
        "call_json": json.dumps(data)
    }

    return render(request, template_name, context)

def create_user_view(request):
    if request.method == "POST":

        username = request.POST.get("username")
        tenant_id = request.POST.get("tenant_id")
        branch_code = request.POST.get("branch_code")

        if request.user.is_authenticated and request.user.role == "admin":

            # ===============================
            # 2️⃣ CHECK TENANT ALREADY EXISTS
            # ===============================
            if Tenant.objects.filter(tenant_id=tenant_id).exists():
                return render(request, "clients.html", {
                    "error": "Tenant ID already exists!"
                })
            


            # ===============================
            # 4️⃣ CREATE TENANT
            # ===============================
            tenant = Tenant.objects.create(
                owner=request.user,
                tenant_id=tenant_id,
                tenant_name=request.POST.get("tenant_name"),
                email=request.POST.get("tenant_email"),
                phone=request.POST.get("tenant_phone"),
                fax=request.POST.get("tenant_fax"),
                address=request.POST.get("tenant_address"),
                city=request.POST.get("tenant_city"),
                state=request.POST.get("tenant_state"),
                country=request.POST.get("tenant_country"),
                postcode=request.POST.get("tenant_postcode"),
            )

            # ===============================
            # 5️⃣ CHECK BRANCH EXISTS (Inside Tenant)
            # ===============================
            if Branch.objects.filter(
                tenant=tenant,
                branch_code=branch_code
            ).exists():
                return render(request, "your_template.html", {
                    "error": "Branch code already exists for this tenant!"
                })

            # ===============================
            # 6️⃣ CREATE BRANCH
            # ===============================
            branch = Branch.objects.create(
                tenant=tenant,
                name=request.POST.get("branch_name"),
                branch_code=branch_code,
                classification=request.POST.get("branch_classification"),
                start_date=request.POST.get("start_date"),
                end_date=request.POST.get("end_date"),
                plan_type=request.POST.get("plan_type"),
                email=request.POST.get("branch_email"),
                phone=request.POST.get("branch_phone"),
                fax=request.POST.get("branch_fax"),
                address=request.POST.get("branch_address"),
                city=request.POST.get("branch_city"),
                state=request.POST.get("branch_state"),
                country=request.POST.get("branch_country"),
                postcode=request.POST.get("branch_postcode"),
            )

            # ===============================
            # 7️⃣ CREATE DEPARTMENT
            # ===============================
            # Department.objects.create(
            #     branch=branch,
            #     name=request.POST.get("dept_name"),
            #     code=request.POST.get("dept_code"),
            #     email=request.POST.get("dept_email"),
            #     phone=request.POST.get("dept_phone"),
            #     fax=request.POST.get("dept_fax"),
            #     address=request.POST.get("dept_address"),
            #     city=request.POST.get("dept_city"),
            #     postcode=request.POST.get("dept_postcode"),
            #     state=request.POST.get("dept_state"),
            #     country=request.POST.get("dept_country"),
            # )

            # ===============================
            # 8️⃣ CREATE WAREHOUSE
            # ===============================
            # Warehouse.objects.create(
            #     branch=branch,
            #     warehouse_id=request.POST.get("warehouse_id"),
            #     name=request.POST.get("warehouse_name"),
            #     address=request.POST.get("warehouse_address"),
            #     city=request.POST.get("warehouse_city"),
            #     postcode=request.POST.get("warehouse_Postcode"),
            #     state=request.POST.get("warehouse__state"),
            #     country=request.POST.get("warehouse_country"),
            # )

            return redirect("client")

        # ===============================
        # 1️⃣ CHECK USER ALREADY EXISTS
        # ===============================
        if CustomUser.objects.filter(username=username).exists():
            return render(request, "clients.html", {
                "error": "Username already exists!"
            })

        # ===============================
        # 2️⃣ CHECK TENANT ALREADY EXISTS
        # ===============================
        if Tenant.objects.filter(tenant_id=tenant_id).exists():
            return render(request, "clients.html", {
                "error": "Tenant ID already exists!"
            })
        

        # ===============================
        # 3️⃣ CREATE USER
        # ===============================
        user = CustomUser.objects.create_user(
            username=username,
            password=request.POST.get("password"),
            role=request.POST.get("role")
        )

        # ===============================
        # 4️⃣ CREATE TENANT
        # ===============================
        tenant = Tenant.objects.create(
            owner=user,
            tenant_id=tenant_id,
            tenant_name=request.POST.get("tenant_name"),
            email=request.POST.get("tenant_email"),
            phone=request.POST.get("tenant_phone"),
            fax=request.POST.get("tenant_fax"),
            address=request.POST.get("tenant_address"),
            city=request.POST.get("tenant_city"),
            state=request.POST.get("tenant_state"),
            country=request.POST.get("tenant_country"),
            postcode=request.POST.get("tenant_postcode"),
        )

        # ===============================
        # 5️⃣ CHECK BRANCH EXISTS (Inside Tenant)
        # ===============================
        if Branch.objects.filter(
            tenant=tenant,
            branch_code=branch_code
        ).exists():
            return render(request, "your_template.html", {
                "error": "Branch code already exists for this tenant!"
            })

        # ===============================
        # 6️⃣ CREATE BRANCH
        # ===============================
        branch = Branch.objects.create(
            tenant=tenant,
            name=request.POST.get("branch_name"),
            branch_code=branch_code,
            classification=request.POST.get("branch_classification"),
            start_date=request.POST.get("start_date"),
            end_date=request.POST.get("end_date"),
            plan_type=request.POST.get("plan_type"),
            email=request.POST.get("branch_email"),
            phone=request.POST.get("branch_phone"),
            fax=request.POST.get("branch_fax"),
            address=request.POST.get("branch_address"),
            city=request.POST.get("branch_city"),
            state=request.POST.get("branch_state"),
            country=request.POST.get("branch_country"),
            postcode=request.POST.get("branch_postcode"),
        )

        # ===============================
        # 7️⃣ CREATE DEPARTMENT
        # ===============================
        Department.objects.create(
            branch=branch,
            name=request.POST.get("dept_name"),
            code=request.POST.get("dept_code"),
            email=request.POST.get("dept_email"),
            phone=request.POST.get("dept_phone"),
            fax=request.POST.get("dept_fax"),
            address=request.POST.get("dept_address"),
            city=request.POST.get("dept_city"),
            postcode=request.POST.get("dept_postcode"),
            state=request.POST.get("dept_state"),
            country=request.POST.get("dept_country"),
        )

        # ===============================
        # 8️⃣ CREATE WAREHOUSE
        # ===============================
        Warehouse.objects.create(
            branch=branch,
            warehouse_id=request.POST.get("warehouse_id"),
            name=request.POST.get("warehouse_name"),
            address=request.POST.get("warehouse_address"),
            city=request.POST.get("warehouse_city"),
            postcode=request.POST.get("warehouse_Postcode"),
            state=request.POST.get("warehouse__state"),
            country=request.POST.get("warehouse_country"),
        )

        return redirect("client")
    return redirect("client")

def branch(request):

# 🔥 Super Admin sees all
    if not request.user.is_authenticated:
        return redirect('login') 
    
    if request.user.is_superuser:
        branches = Branch.objects.all()
    else:
        # Normal user sees only their branches
        branches = Branch.objects.filter(
            tenant__owner=request.user
        )


    data = []
    tenants_drop = Tenant.objects.all()
    for b in branches:
        data.append({
            "id": b.id,
            "name": b.name,
            "branch_code": b.branch_code,
            "classification": b.classification,
            "start_date": b.start_date.strftime("%Y-%m-%d") if b.start_date else "",
            "end_date": b.end_date.strftime("%Y-%m-%d") if b.end_date else "",
            "plan_type": b.plan_type,

            "tenant_name": b.tenant.tenant_name if b.tenant else "",
            "tenant_id": b.tenant.tenant_id if b.tenant else "",

            "brn": b.brn,
            "sst_number": b.sst_number,
            "industry_code": b.industry_code,
            "business_activity": b.business_activity,

            
            "clinic_header": b.clinic_header,
            "controlled_medicine": b.controlled_medicine,
            "dental_module": b.dental_module,

            "date": b.created_at.strftime("%Y-%m-%d") if b.created_at else None,

            "email": b.email,
            "phone": b.phone,
            "fax": b.fax,

            "address": b.address,
            "city": b.city,
            "state": b.state,
            "country": b.country,
            "postcode": b.postcode,
        })
    print("data :", data)

    return render(request, "admin_branch.html", {
        "call_json": json.dumps(data), 
        "tenants_drop":tenants_drop
    })

def branch_api(request):

    if request.user.is_superuser:
        branches = Branch.objects.select_related("tenant").all()
    else:
        branches = Branch.objects.select_related("tenant").filter(
            tenant__owner=request.user
        )

    search = request.GET.get("search", "")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    page = int(request.GET.get("page", 1))

    if search:
        branches = branches.filter(name__icontains=search)

    if from_date:
        branches = branches.filter(created_at__date__gte=from_date)

    if to_date:
        branches = branches.filter(created_at__date__lte=to_date)

    paginator = Paginator(branches, 10)
    page_obj = paginator.get_page(page)

    data = []

    for b in page_obj:
        data.append({
            "id": b.id,
            "name": b.name,
            "branch_code": b.branch_code,

            # 🔥 IMPORTANT PART
            "tenant_name": b.tenant.tenant_name if b.tenant else "",
            "tenant_id": b.tenant.tenant_id if b.tenant else "",

            "classification": b.classification,
            "start_date": b.start_date.strftime("%Y-%m-%d") if b.start_date else "",
            "end_date": b.end_date.strftime("%Y-%m-%d") if b.end_date else "",
            "plan_type": b.plan_type,

            "date": b.created_at.strftime("%Y-%m-%d") if b.created_at else "",

            "city": b.city,
            "email": b.email,
            "phone": b.phone,
            "fax": b.fax,
            "address": b.address,
            "state": b.state,
            "country": b.country,
            "postcode": b.postcode,

            "brn": b.brn,
            "sst_number": b.sst_number,
            "industry_code": b.industry_code,
            "business_activity": b.business_activity,

            "clinic_header": b.clinic_header,
            "controlled_medicine": b.controlled_medicine,
            "dental_module": b.dental_module,
        })

    return JsonResponse({
        "data": data,
        "total_pages": paginator.num_pages
    })
def create_branch(request):
    if request.method == "POST":
        tenant = Tenant.objects.get(id= request.POST.get("branch_tenant"))
        branch_code = request.POST.get("branch_code")
        if Branch.objects.filter(branch_code=branch_code).exists():
            messages.error(request, "Branch code already exists!")
            return redirect("branch")  # your branch page name
        
        branch = Branch.objects.create(
        tenant=tenant,
        name=request.POST.get("branch_name"),
        branch_code=request.POST.get("branch_code"),
        classification=request.POST.get("branch_classification"),
        start_date=request.POST.get("start_date"),
        end_date=request.POST.get("end_date"),
        plan_type=request.POST.get("plan_type"),

        brn=request.POST.get("brn"),
        sst_number=request.POST.get("sst_number"),
        # industry_code=request.POST.get("industry_code"),
        # business_activity=request.POST.get("business_activity"),

        # clinic_header=request.POST.get("clinic_header") == "True",
        # controlled_medicine=request.POST.get("controlled_medicine") == "True",
        # dental_module=request.POST.get("dental_module") == "True",

        # mi2u_expiry=request.POST.get("mi2u_expiry"),

        email=request.POST.get("branch_email"),
        phone=request.POST.get("branch_phone"),
        fax=request.POST.get("branch_fax"),
        address=request.POST.get("branch_address"),
        city=request.POST.get("branch_city"),
        state=request.POST.get("branch_state"),
        country=request.POST.get("branch_country"),
        postcode=request.POST.get("branch_postcode"),
        
        )
        messages.success(request, "Branch created successfully!")
        return redirect("branch")
    return render(request, "admin_branch.html")

def assistant(request):
    if request.user.is_superuser:
        branches = Branch.objects.select_related("tenant").all()
        
    else:
        # ✅ Only current logged-in user's tenant branches
        branches = Branch.objects.select_related("tenant").filter(
            tenant__owner=request.user
        )
        

    data = []

    for b in branches:
        data.append({
            "tenant_id": b.tenant.tenant_id,
            "tenant_name": b.tenant.tenant_name,
            "branch_id": b.branch_code,
            "branch_name": b.name,
        })

    return render(request, "Assistant.html", {"data":data})

def create_ai_assistant(request):

    if request.method == "POST":

        agent_name = request.POST.get("agent_name")
        primary_language = request.POST.get("primary_language")
        secondary_languages = request.POST.getlist("secondary_languages")
        system_prompt = request.POST.get("system_prompt")
        tenant_id = request.POST.get("tenant_id")
        branch_id = request.POST.get("branch_id")
        voice_name = request.POST.get("voice_name")

        uploaded_files = request.FILES.getlist("uploaded_files")

        time.sleep(5)

        for file in uploaded_files:

            AIAssistant.objects.create(
                agent_name=agent_name,
                primary_language=primary_language,
                secondary_languages=", ".join(secondary_languages),
                system_prompt=system_prompt,
                tenant_id=tenant_id,
                branch_id=branch_id,
                voice_name=voice_name,
                uploaded_file=file,
                created_by=request.user
            )

        return render(request, "loader.html")

    return redirect("assistant")

def loader_page(request):
    return render(request, "loader.html")

def assistant_list(request):

    if request.user.is_superuser:
        assistants = AIAssistant.objects.all().order_by("-created_at")
    else:
        assistants = AIAssistant.objects.filter(
            created_by=request.user
        ).order_by("-created_at")

    return render(request, "assistant_list.html", {
        "assistants": assistants
    })

def widgetconfiguration(request):
    return render(request, "widgetconfiguration.html")

def widget_configuration(request, pk):
    assistant = AIAssistant.objects.get(id=pk)
    return render(request, "widgetconfiguration.html", {
        "assistant": assistant
    })


@csrf_exempt
def delete_selected_widget_calls(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)
    print("this delete .....")
    try:
        print("try")
        body = json.loads(request.body)
        ids = body.get("ids", [])
        print(ids,"ids")

        if not ids:
            return JsonResponse({"deleted": 0})

        deleted_count, _ = ElevenCall.objects.filter(id__in=ids).delete()

        return JsonResponse({
            "deleted": deleted_count
        })

    except Exception as e:
        print("errorv : ", e)
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def delete_selected_tenants(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)
    print("this delete .....")
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Unauthorized"}, status=401)
    try:
        body = json.loads(request.body)
        ids = body.get("ids", [])

        deleted_count, _ = Tenant.objects.filter(id__in=ids).delete()

        return JsonResponse({"deleted": deleted_count})
    except Exception as e:
        print("errorv : ", e)
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def delete_selected_branches(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)
    print("this delete .....")
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Unauthorized"}, status=401)
    try:
        body = json.loads(request.body)
        ids = body.get("ids", [])

        deleted_count, _ = Branch.objects.filter(id__in=ids).delete()

        return JsonResponse({"deleted": deleted_count})
    except Exception as e:
        print("errorv : ", e)
        return JsonResponse({"error": str(e)}, status=500)

def otp_password(request):
    data = "8811"
    return JsonResponse({"data":data}) 